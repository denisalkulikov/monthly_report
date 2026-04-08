from nicegui import ui
import psycopg2
from dotenv import load_dotenv
import os
import warnings
import openpyxl
from openpyxl import load_workbook
import asyncio
import tempfile
import traceback
import shutil

load_dotenv()

warnings.filterwarnings('ignore')

# Конфигурация БД из переменных окружения
DB_CONFIG = {
    "host": os.getenv("DB_HOST", "localhost"),
    "database": os.getenv("DB_NAME", "postgres"),
    "user": os.getenv("DB_USER", "postgres"),
    "password": os.getenv("DB_PASSWORD", ""),
    "port": os.getenv("DB_PORT", "5432")
}

# Проверка конфигурации
print("Конфигурация БД:")
for key, value in DB_CONFIG.items():
    if key == 'password':
        print(f"  {key}: {'*' * len(value)}")
    else:
        print(f"  {key}: {value}")

year = [
    2020, 2021, 2022, 2023, 2024, 2025, 2026
]

month = [
    'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
    'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'
]

# Словарь для преобразования названия месяца в номер
month_to_number = {
    'Январь': 1, 'Февраль': 2, 'Март': 3, 'Апрель': 4,
    'Май': 5, 'Июнь': 6, 'Июль': 7, 'Август': 8,
    'Сентябрь': 9, 'Октябрь': 10, 'Ноябрь': 11, 'Декабрь': 12
}

# Словарь для преобразования названия месяца в именительный падеж с маленькой буквы
month_to_nominative_lower = {
    'Январь': 'январь',
    'Февраль': 'февраль',
    'Март': 'март',
    'Апрель': 'апрель',
    'Май': 'май',
    'Июнь': 'июнь',
    'Июль': 'июль',
    'Август': 'август',
    'Сентябрь': 'сентябрь',
    'Октябрь': 'октябрь',
    'Ноябрь': 'ноябрь',
    'Декабрь': 'декабрь'
}

# Словарь для преобразования названия месяца в предложный падеж
month_to_prepositional = {
    'Январь': 'январе',
    'Февраль': 'феврале',
    'Март': 'марте',
    'Апрель': 'апреле',
    'Май': 'мае',
    'Июнь': 'июне',
    'Июль': 'июле',
    'Август': 'августе',
    'Сентябрь': 'сентябре',
    'Октябрь': 'октябре',
    'Ноябрь': 'ноябре',
    'Декабрь': 'декабре'
}

# Словарь для преобразования названия месяца в букву колонки
month_to_column = {
    1: 'B', 2: 'C', 3: 'D', 4: 'E', 5: 'F', 6: 'G',
    7: 'H', 8: 'I', 9: 'J', 10: 'K', 11: 'L', 12: 'M'
}

# Глобальные переменные для хранения данных
cached_data = None
cached_sales_data = None
cached_monthly_group_data = None
cached_sales_responsibility_data = None
cached_tk_monthly_group_data = None
cached_tk_sales_responsibility_data = None
cached_reklama_monthly_group_data = None
cached_reklama_sales_responsibility_data = None
cached_reklama_total = None
cached_angar_monthly_group_data = None
cached_angar_sales_data = None
cached_kn_monthly_group_data = None
cached_kn_sales_data = None
cached_kn_sales_responsibility_data = None
cached_total_shipping_sum_angar = None
cached_total_shipping_sum_kn = None
cached_total_shipping_sum_reklama = None
cached_total_shipping_sum_tk = None
current_results_container = None

# Переменные для хранения файлов
file1_path = None
file2_path = None


def get_db_connection():
    """Создание подключения к БД"""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        return conn
    except Exception as e:
        error_msg = f'Ошибка подключения к БД: {e}'
        print(error_msg)
        print(traceback.format_exc())
        ui.notify(error_msg, type='negative')
        return None


def fetch_total_shipping_sum(year_value, month_number, direction):
    """Запрос для получения общей суммы shipping_sum по направлению"""
    try:
        conn = get_db_connection()
        if conn is None:
            return 0

        cur = conn.cursor()

        query = """
                SELECT COALESCE(SUM(shipping_sum), 0) as total_summ
                FROM kamtent.sales
                WHERE EXTRACT(YEAR FROM TO_DATE(shipping_date, 'YYYY-MM-DD')) = %s
                  AND EXTRACT(MONTH FROM TO_DATE(shipping_date, 'YYYY-MM-DD')) = %s
                  AND direction = %s
                  AND shipping_date IS NOT NULL
                  AND shipping_date != '' \
                """

        print(f"Выполняю запрос общей shipping_sum для {direction}")
        cur.execute(query, (year_value, month_number, direction))
        result = cur.fetchone()

        cur.close()
        conn.close()

        total = result[0] if result and result[0] else 0
        print(f"Общая shipping_sum для {direction}: {total}")

        return total

    except Exception as e:
        error_msg = f'Ошибка при выполнении запроса общей shipping_sum: {e}'
        print(error_msg)
        print(traceback.format_exc())
        ui.notify(error_msg, type='negative')
        return 0


def fetch_data_by_direction(year_value, month_number):
    """Запрос для получения сумм по направлениям"""
    try:
        conn = get_db_connection()
        if conn is None:
            return None

        cur = conn.cursor()

        query = """
                SELECT direction, SUM(pay_summ) as total_summ
                FROM kamtent.monthly_group_product
                WHERE year = %s \
                  AND month = %s
                GROUP BY direction
                ORDER BY direction \
                """

        cur.execute(query, (year_value, month_number))
        results = cur.fetchall()

        cur.close()
        conn.close()

        return results

    except Exception as e:
        error_msg = f'Ошибка при выполнении запроса: {e}'
        print(error_msg)
        print(traceback.format_exc())
        ui.notify(error_msg, type='negative')
        return None


def fetch_data_for_directions(year_value, month_number):
    """Запрос для получения сумм по конкретным направлениям"""
    try:
        conn = get_db_connection()
        if conn is None:
            return None

        cur = conn.cursor()

        query = """
                SELECT direction, SUM(pay_summ) as total_summ
                FROM kamtent.monthly_group_product
                WHERE year = %s \
                  AND month = %s
                  AND UPPER (direction) IN ('ОАИ' \
                    , 'ТК' \
                    , 'АНГАРЫ' \
                    , 'РЕКЛАМА' \
                    , 'КН')
                GROUP BY direction \
                """

        cur.execute(query, (year_value, month_number))
        results = cur.fetchall()

        cur.close()
        conn.close()

        result_dict = {row[0]: row[1] for row in results}
        print(f"Данные из БД для направлений: {result_dict}")

        return result_dict

    except Exception as e:
        error_msg = f'Ошибка при выполнении запроса: {e}'
        print(error_msg)
        print(traceback.format_exc())
        ui.notify(error_msg, type='negative')
        return None


def fetch_sales_data(year_value, month_number):
    """Запрос для получения данных из таблицы sales для листа Авто"""
    try:
        conn = get_db_connection()
        if conn is None:
            return None

        cur = conn.cursor()

        query = """
                SELECT segment, SUM(pay_summ) as total_summ
                FROM kamtent.sales
                WHERE EXTRACT(YEAR FROM pay_date) = %s
                  AND EXTRACT(MONTH FROM pay_date) = %s
                  AND direction = 'ОАИ'
                  AND segment IN ('РОЗНИЦА', 'ПОТРЕБИТЕЛИ')
                  AND pay_date IS NOT NULL
                GROUP BY segment \
                """

        cur.execute(query, (year_value, month_number))
        results = cur.fetchall()

        cur.close()
        conn.close()

        result_dict = {row[0]: row[1] for row in results}
        print(f"Данные из sales для segment: {result_dict}")

        return result_dict

    except Exception as e:
        error_msg = f'Ошибка при выполнении запроса к sales: {e}'
        print(error_msg)
        print(traceback.format_exc())
        ui.notify(error_msg, type='negative')
        return None


def fetch_kn_sales_data(year_value, month_number):
    """Запрос для получения данных из таблицы sales для листа КН"""
    try:
        conn = get_db_connection()
        if conn is None:
            return None

        cur = conn.cursor()

        query = """
                SELECT segment, SUM(pay_summ) as total_summ
                FROM kamtent.sales
                WHERE EXTRACT(YEAR FROM pay_date) = %s
                  AND EXTRACT(MONTH FROM pay_date) = %s
                  AND direction = 'КН'
                  AND segment IN ('РОЗНИЦА', 'ПОТРЕБИТЕЛИ')
                  AND pay_date IS NOT NULL
                GROUP BY segment \
                """

        cur.execute(query, (year_value, month_number))
        results = cur.fetchall()

        cur.close()
        conn.close()

        result_dict = {row[0]: row[1] for row in results}
        print(f"Данные из sales для segment (КН): {result_dict}")

        return result_dict

    except Exception as e:
        error_msg = f'Ошибка при выполнении запроса к sales (КН): {e}'
        print(error_msg)
        print(traceback.format_exc())
        ui.notify(error_msg, type='negative')
        return None


def fetch_angar_sales_data(year_value, month_number):
    """Запрос для получения данных из таблицы sales для листа Ангар"""
    try:
        conn = get_db_connection()
        if conn is None:
            return None

        cur = conn.cursor()

        query = """
                SELECT group_product, SUM(shipping_sum) as total_summ
                FROM kamtent.sales
                WHERE EXTRACT(YEAR FROM TO_DATE(shipping_date, 'YYYY-MM-DD')) = %s
                  AND EXTRACT(MONTH FROM TO_DATE(shipping_date, 'YYYY-MM-DD')) = %s
                  AND direction = 'АНГАРЫ'
                  AND group_product IN ('СПОРТ И КУЛЬТУРА', 'ПРОЧЕЕ', 'СЕЛЬСКОЕ ХОЗЯЙСТВО', 'ПРОМЫШЛЕННОСТЬ')
                  AND shipping_date IS NOT NULL
                  AND shipping_date != ''
                GROUP BY group_product \
                """

        cur.execute(query, (year_value, month_number))
        results = cur.fetchall()

        cur.close()
        conn.close()

        result_dict = {row[0]: row[1] for row in results}
        print(f"Данные из sales для Ангар: {result_dict}")

        return result_dict

    except Exception as e:
        error_msg = f'Ошибка при выполнении запроса к sales (Ангар): {e}'
        print(error_msg)
        print(traceback.format_exc())
        ui.notify(error_msg, type='negative')
        return None


def fetch_sales_responsibility_data(year_value, month_number, direction='ОАИ'):
    """Запрос для получения данных из таблицы sales по responsibility"""
    try:
        conn = get_db_connection()
        if conn is None:
            return None

        cur = conn.cursor()

        query = """
                SELECT responsibility, SUM(pay_summ) as total_summ
                FROM kamtent.sales
                WHERE EXTRACT(YEAR FROM pay_date) = %s
                  AND EXTRACT(MONTH FROM pay_date) = %s
                  AND direction = %s
                  AND LOWER(responsibility) IN ('своя', 'чужая')
                  AND pay_date IS NOT NULL
                GROUP BY responsibility \
                """

        cur.execute(query, (year_value, month_number, direction))
        results = cur.fetchall()

        cur.close()
        conn.close()

        result_dict = {row[0].upper(): row[1] for row in results}
        print(f"Данные из sales для responsibility (direction={direction}): {result_dict}")

        return result_dict

    except Exception as e:
        error_msg = f'Ошибка при выполнении запроса к sales (responsibility): {e}'
        print(error_msg)
        print(traceback.format_exc())
        ui.notify(error_msg, type='negative')
        return None


def fetch_monthly_group_products(year_value, month_number, direction='ОАИ'):
    """Запрос для получения данных из monthly_group_product"""
    try:
        conn = get_db_connection()
        if conn is None:
            return None

        cur = conn.cursor()

        if direction == 'ОАИ':
            group_products = ['МСК', 'АВТОТЕНТЫ', 'АВТОУСЛУГИ', 'РЕМОНТ', 'ПРОЧЕЕ', 'АВТОКАРКАСЫ', 'ВОРОТА',
                              'АВТОПОЛОГИ']
        elif direction == 'ТК':
            group_products = ['ОРИГИНАЛЬНЫЕ ТК', 'ПРОМЫШЛЕННЫЕ ТК', 'ПРОЧЕЕ', 'СЕЛЬСКОХОЗЯЙСТВЕННЫЕ ТК',
                              'СПОРТИВНЫЕ И КУЛЬТ. ТК', 'ТОРГОВЫЕ ТК']
        elif direction == 'РЕКЛАМА':
            group_products = ['РЕКЛАМА Т', 'РЕКЛАМА А', 'РЕКЛАМА П', 'РЕКЛАМА Б']
        elif direction == 'АНГАРЫ':
            group_products = ['СПОРТ И КУЛЬТУРА', 'ПРОЧЕЕ', 'СЕЛЬСКОЕ ХОЗЯЙСТВО', 'ПРОМЫШЛЕННОСТЬ']
        elif direction == 'КН':
            group_products = ['ФУРНИТУРА', 'ПРОЧЕЕ', 'ТКАНИ']
        else:
            return {}

        placeholders = ', '.join(['%s'] * len(group_products))
        query = f"""
        SELECT group_product, SUM(pay_summ) as total_summ
        FROM kamtent.monthly_group_product
        WHERE year = %s AND month = %s 
        AND direction = %s
        AND group_product IN ({placeholders})
        GROUP BY group_product
        """

        params = [year_value, month_number, direction] + group_products
        cur.execute(query, params)
        results = cur.fetchall()

        cur.close()
        conn.close()

        result_dict = {row[0]: row[1] for row in results}
        print(f"Данные из monthly_group_product для direction={direction}: {result_dict}")

        return result_dict

    except Exception as e:
        error_msg = f'Ошибка при выполнении запроса к monthly_group_product: {e}'
        print(error_msg)
        print(traceback.format_exc())
        ui.notify(error_msg, type='negative')
        return None


def fetch_reklama_total(year_value, month_number):
    """Запрос для получения общей суммы по направлению РЕКЛАМА"""
    try:
        conn = get_db_connection()
        if conn is None:
            return 0

        cur = conn.cursor()

        query = """
                SELECT COALESCE(SUM(pay_summ), 0) as total_summ
                FROM kamtent.monthly_group_product
                WHERE year = %s \
                  AND month = %s
                  AND direction = 'РЕКЛАМА' \
                """

        cur.execute(query, (year_value, month_number))
        result = cur.fetchone()

        cur.close()
        conn.close()

        total = result[0] if result and result[0] else 0
        print(f"Общая сумма по направлению РЕКЛАМА: {total}")

        return total

    except Exception as e:
        error_msg = f'Ошибка при выполнении запроса к monthly_group_product (total): {e}'
        print(error_msg)
        print(traceback.format_exc())
        ui.notify(error_msg, type='negative')
        return 0


def fetch_group_products(year_value, month_number, direction):
    """Запрос для получения group_product по конкретному направлению"""
    try:
        conn = get_db_connection()
        if conn is None:
            return None

        cur = conn.cursor()

        query = """
                SELECT group_product, pay_summ
                FROM kamtent.monthly_group_product
                WHERE year = %s \
                  AND month = %s \
                  AND direction = %s
                ORDER BY group_product \
                """

        cur.execute(query, (year_value, month_number, direction))
        results = cur.fetchall()

        cur.close()
        conn.close()

        return results

    except Exception as e:
        error_msg = f'Ошибка при выполнении запроса: {e}'
        print(error_msg)
        print(traceback.format_exc())
        ui.notify(error_msg, type='negative')
        return None


def create_expandable_row(direction, direction_total, year_value, month_number):
    """Создает раскрывающуюся строку для направления"""
    with ui.expansion(f'{direction} — {direction_total:,.2f} руб.', icon='folder').classes('w-full mb-2'):
        content_container = ui.column()

        with content_container:
            with ui.row().classes('items-center gap-2'):
                ui.spinner(size='sm')
                ui.label('Загрузка данных...').classes('text-grey-6')

        group_products = fetch_group_products(year_value, month_number, direction)

        content_container.clear()

        with content_container:
            if group_products and len(group_products) > 0:
                with ui.row().classes('justify-between items-center mb-2 p-2 bg-grey-1 rounded'):
                    ui.label('Общая сумма направления:').classes('text-subtitle1 font-bold')
                    ui.label(f'{direction_total:,.2f} руб.').classes('text-subtitle1 font-bold text-primary')

                columns = [
                    {'name': 'group_product', 'label': 'Группа продуктов', 'field': 'group_product', 'align': 'left'},
                    {'name': 'pay_summ', 'label': 'Сумма, руб.', 'field': 'pay_summ', 'align': 'right'}
                ]

                rows = []
                subtotal = 0

                for product, summ in group_products:
                    if summ:
                        rows.append({'group_product': product, 'pay_summ': f'{float(summ):,.2f}'})
                        subtotal += float(summ)
                    else:
                        rows.append({'group_product': product, 'pay_summ': '0.00'})

                ui.table(columns=columns, rows=rows, row_key='group_product').classes('w-full')

                with ui.row().classes('mt-2 justify-end w-full'):
                    ui.label(f'Сумма по группам:').classes('text-subtitle2')
                    ui.label(f'{subtotal:,.2f} руб.').classes('text-subtitle2 text-primary')

                if abs(subtotal - direction_total) > 0.01:
                    with ui.row().classes('mt-1 justify-end w-full'):
                        ui.label(f'(Расхождение: {direction_total - subtotal:,.2f} руб.)').classes(
                            'text-caption text-grey-6')
            else:
                ui.label('Нет данных по продуктам в этом направлении').classes('text-grey-6')


def copy_cells_between_files(source_path, target_path):
    """Копирование значений ячеек из второго файла в первый"""
    try:
        print(f"Начинаю копирование из {source_path} в {target_path}")

        # Загружаем оба файла
        source_wb = load_workbook(source_path, keep_vba=True)
        target_wb = load_workbook(target_path, keep_vba=True)

        # Сопоставление листов и ячеек для копирования
        copy_mapping = {
            'Ателье': {
                'source': ['Q31', 'Q32', 'Q33'],
                'target': ['P31', 'P32', 'P33']
            },
            'Авто': {
                'source': ['R5', 'R7', 'R36', 'R37', 'R38', 'R39', 'R40', 'R41', 'R42', 'R43', 'R44', 'R45', 'R74',
                           'R75'],
                'target': ['Q5', 'Q7', 'Q36', 'Q37', 'Q38', 'Q39', 'Q40', 'Q41', 'Q42', 'Q43', 'Q44', 'Q45', 'Q74',
                           'Q75']
            },
            'ТК': {
                'source': ['R6', 'R7', 'R8', 'R9', 'R10', 'R11', 'R39', 'R40', 'R41', 'R42', 'R43', 'R44', 'R45', 'R46',
                           'R47', 'R87', 'R88'],
                'target': ['Q6', 'Q7', 'Q8', 'Q9', 'Q10', 'Q11', 'Q39', 'Q40', 'Q41', 'Q42', 'Q43', 'Q44', 'Q45', 'Q46',
                           'Q47', 'Q87', 'Q88']
            },
            'Реклама': {
                'source': ['R6', 'R7', 'R8', 'R9', 'R10', 'R11', 'R39', 'R40', 'R41', 'R42', 'R43', 'R70', 'R71'],
                'target': ['Q6', 'Q7', 'Q8', 'Q9', 'Q10', 'Q11', 'Q39', 'Q40', 'Q41', 'Q42', 'Q43', 'Q70', 'Q71']
            },
            'Ангар': {
                'source': ['Q6', 'Q7', 'Q8', 'Q9', 'Q33', 'Q34', 'Q35', 'Q36'],
                'target': ['P6', 'P7', 'P8', 'P9', 'P33', 'P34', 'P35', 'P36']
            },
            'Ком': {
                'source': ['R6', 'R7', 'R8', 'R29', 'R31', 'R58', 'R59'],
                'target': ['Q6', 'Q7', 'Q8', 'Q29', 'Q31', 'Q58', 'Q59']
            }
        }

        # Копируем значения
        for sheet_name, cells in copy_mapping.items():
            if sheet_name in source_wb.sheetnames and sheet_name in target_wb.sheetnames:
                source_sheet = source_wb[sheet_name]
                target_sheet = target_wb[sheet_name]

                print(f"Копирую с листа {sheet_name}")

                for src_cell, tgt_cell in zip(cells['source'], cells['target']):
                    value = source_sheet[src_cell].value
                    target_sheet[tgt_cell] = value
                    print(f"  {src_cell} -> {tgt_cell}: {value}")
            else:
                print(f"Лист {sheet_name} не найден в одном из файлов")

        # Сохраняем целевой файл
        output_filename = os.path.basename(target_path)
        output_path = os.path.join(os.getcwd(), output_filename)
        target_wb.save(output_path)
        target_wb.close()
        source_wb.close()

        print(f"Файл успешно сохранен как: {output_path}")
        return output_path

    except Exception as e:
        error_msg = f'Ошибка при копировании: {e}'
        print(error_msg)
        print(traceback.format_exc())
        ui.notify(error_msg, type='negative')
        return None


def process_excel_file(file1_path, file2_path, selected_month_name, selected_year,
                       directions_data, sales_data, monthly_group_data,
                       sales_responsibility_data, tk_monthly_group_data,
                       tk_sales_responsibility_data, reklama_monthly_group_data,
                       reklama_sales_responsibility_data, reklama_total,
                       angar_monthly_group_data, angar_sales_data,
                       kn_monthly_group_data, kn_sales_data, kn_sales_responsibility_data,
                       total_shipping_angar, total_shipping_kn, total_shipping_reklama,
                       total_shipping_tk, realization_amount_value, client, progress_callback=None):
    """Обработка Excel файла"""
    output_path = None
    temp_file_path = None
    try:
        # Создаем временную копию первого файла
        if progress_callback:
            progress_callback(10, "Копирование файла-шаблона...")
        temp_file_path = tempfile.mktemp(suffix='.xlsm')
        shutil.copy2(file1_path, temp_file_path)

        # Загружаем workbook из временного файла
        if progress_callback:
            progress_callback(20, "Загрузка файла...")
        wb = load_workbook(temp_file_path, keep_vba=True)

        # 1. Обработка листа "служ"
        if progress_callback:
            progress_callback(25, "Обработка листа 'служ'...")
        if "служ" in wb.sheetnames:
            sheet_serv = wb["служ"]
            month_number = month_to_number[selected_month_name]
            month_nominative = month_to_nominative_lower[selected_month_name]
            month_prepositional = month_to_prepositional[selected_month_name]

            sheet_serv['B1'] = f"{month_nominative} "
            sheet_serv['B2'] = f"{month_prepositional} "
            sheet_serv['B3'] = str(month_number)
            sheet_serv['B4'] = str(selected_year)
        else:
            with client:
                ui.notify('Лист "служ" не найден в файле', type='warning')

        # 2. Обработка листа "тит"
        if progress_callback:
            progress_callback(30, "Обработка листа 'тит'...")
        if "тит" in wb.sheetnames:
            sheet_tit = wb["тит"]
            sheet_tit['F44'] = str(selected_year)
        else:
            with client:
                ui.notify('Лист "тит" не найден в файле', type='warning')

        # 3. Обработка листа "общ"
        if progress_callback:
            progress_callback(35, "Обработка листа 'общ'...")
        if "общ" in wb.sheetnames:
            sheet_obsh = wb["общ"]
            month_number = month_to_number[selected_month_name]
            column_letter = month_to_column[month_number]

            direction_cells = {'ОАИ': '40', 'ТК': '41', 'АНГАРЫ': '43', 'РЕКЛАМА': '47', 'КН': '49'}

            for direction, row in direction_cells.items():
                cell = f"{column_letter}{row}"
                amount = 0
                for db_direction, db_amount in directions_data.items():
                    if db_direction.upper() == direction:
                        amount = float(db_amount) if db_amount else 0
                        break
                sheet_obsh[cell] = amount
                print(f"Записываю в ячейку {cell} сумму: {amount}")
        else:
            with client:
                ui.notify('Лист "общ" не найден в файле', type='warning')

        # 4. Обработка листа "сегм"
        if progress_callback:
            progress_callback(45, "Обработка листа 'сегм'...")
        if "сегм" in wb.sheetnames:
            sheet_segm = wb["сегм"]
            month_number = month_to_number[selected_month_name]
            column_letter = month_to_column[month_number]

            cell_angar = f"{column_letter}48"
            sheet_segm[cell_angar] = float(total_shipping_angar) if total_shipping_angar else 0
            print(f"Записываю в ячейку {cell_angar} (shipping_sum АНГАРЫ) сумму: {total_shipping_angar}")

            cell_kn = f"{column_letter}49"
            sheet_segm[cell_kn] = float(total_shipping_kn) if total_shipping_kn else 0
            print(f"Записываю в ячейку {cell_kn} (shipping_sum КН) сумму: {total_shipping_kn}")

            difference = float(realization_amount_value) - float(total_shipping_angar) - float(total_shipping_kn)
            cell_diff = f"{column_letter}47"
            sheet_segm[cell_diff] = difference if difference else 0
            print(f"Записываю в ячейку {cell_diff} (разница) сумму: {difference}")
        else:
            with client:
                ui.notify('Лист "сегм" не найден в файле', type='warning')

        # 5. Обработка листа "Ателье"
        if progress_callback:
            progress_callback(55, "Обработка листа 'Ателье'...")
        if "Ателье" in wb.sheetnames:
            sheet_atelier = wb["Ателье"]

            sheet_atelier['Q33'] = float(total_shipping_reklama) if total_shipping_reklama else 0
            print(f"Записываю в ячейку Q33 (shipping_sum РЕКЛАМА) сумму: {total_shipping_reklama}")

            sheet_atelier['Q32'] = float(total_shipping_tk) if total_shipping_tk else 0
            print(f"Записываю в ячейку Q32 (shipping_sum ТК) сумму: {total_shipping_tk}")

            diff_from_segm = float(realization_amount_value) - float(total_shipping_angar) - float(total_shipping_kn)
            result = diff_from_segm - float(total_shipping_reklama) - float(total_shipping_tk)
            sheet_atelier['Q31'] = result if result else 0
            print(f"Записываю в ячейку Q31 (результат) сумму: {result}")
        else:
            with client:
                ui.notify('Лист "Ателье" не найден в файле', type='warning')

        # 6. Обработка листа "Авто"
        if progress_callback:
            progress_callback(65, "Обработка листа 'Авто'...")
        if "Авто" in wb.sheetnames:
            sheet_auto = wb["Авто"]

            roznica_amount = sales_data.get('РОЗНИЦА', 0)
            potrebiteli_amount = sales_data.get('ПОТРЕБИТЕЛИ', 0)

            sheet_auto['R5'] = float(roznica_amount) if roznica_amount else 0
            sheet_auto['R7'] = float(potrebiteli_amount) if potrebiteli_amount else 0

            group_product_cells = {
                'АВТОТЕНТЫ': 'R36', 'АВТОПОЛОГИ': 'R37', 'АВТОКАРКАСЫ': 'R38',
                'ВОРОТА': 'R39', 'АВТОУСЛУГИ': 'R42', 'МСК': 'R43', 'РЕМОНТ': 'R44', 'ПРОЧЕЕ': 'R45'
            }

            for group_product, cell in group_product_cells.items():
                amount = monthly_group_data.get(group_product, 0)
                sheet_auto[cell] = float(amount) if amount else 0

            svoya_amount = sales_responsibility_data.get('СВОЯ', 0)
            chuzhaya_amount = sales_responsibility_data.get('ЧУЖАЯ', 0)

            sheet_auto['R74'] = float(svoya_amount) if svoya_amount else 0
            sheet_auto['R75'] = float(chuzhaya_amount) if chuzhaya_amount else 0
        else:
            with client:
                ui.notify('Лист "Авто" не найден в файле', type='warning')

        # 7. Обработка листа "ТК"
        if progress_callback:
            progress_callback(75, "Обработка листа 'ТК'...")
        if "ТК" in wb.sheetnames:
            sheet_tk = wb["ТК"]

            tk_group_cells_1 = {
                'ТОРГОВЫЕ ТК': 'R6', 'ПРОМЫШЛЕННЫЕ ТК': 'R7', 'ОРИГИНАЛЬНЫЕ ТК': 'R8',
                'СПОРТИВНЫЕ И КУЛЬТ. ТК': 'R9', 'СЕЛЬСКОХОЗЯЙСТВЕННЫЕ ТК': 'R10', 'ПРОЧЕЕ': 'R11'
            }

            for group_product, cell in tk_group_cells_1.items():
                amount = tk_monthly_group_data.get(group_product, 0)
                sheet_tk[cell] = float(amount) if amount else 0

            tk_group_cells_2 = {
                'ТОРГОВЫЕ ТК': 'R40', 'СПОРТИВНЫЕ И КУЛЬТ. ТК': 'R41',
                'ПРОМЫШЛЕННЫЕ ТК': 'R42', 'СЕЛЬСКОХОЗЯЙСТВЕННЫЕ ТК': 'R43'
            }

            for group_product, cell in tk_group_cells_2.items():
                amount = tk_monthly_group_data.get(group_product, 0)
                sheet_tk[cell] = float(amount) if amount else 0

            original_amount = tk_monthly_group_data.get('ОРИГИНАЛЬНЫЕ ТК', 0)
            prochee_amount = tk_monthly_group_data.get('ПРОЧЕЕ', 0)
            sheet_tk['R46'] = float(original_amount) + float(prochee_amount)

            tk_svoya_amount = tk_sales_responsibility_data.get('СВОЯ', 0)
            tk_chuzhaya_amount = tk_sales_responsibility_data.get('ЧУЖАЯ', 0)

            sheet_tk['R87'] = float(tk_svoya_amount) if tk_svoya_amount else 0
            sheet_tk['R88'] = float(tk_chuzhaya_amount) if tk_chuzhaya_amount else 0
        else:
            with client:
                ui.notify('Лист "ТК" не найден в файле', type='warning')

        # 8. Обработка листа "Реклама"
        if progress_callback:
            progress_callback(85, "Обработка листа 'Реклама'...")
        if "Реклама" in wb.sheetnames:
            sheet_reklama = wb["Реклама"]

            reklama_cells = {
                'РЕКЛАМА П': 'R6', 'РЕКЛАМА Б': 'R7', 'РЕКЛАМА Т': 'R9', 'РЕКЛАМА А': 'R10'
            }

            for group_product, cell in reklama_cells.items():
                amount = reklama_monthly_group_data.get(group_product, 0)
                sheet_reklama[cell] = float(amount) if amount else 0

            sheet_reklama['R39'] = float(reklama_total) if reklama_total else 0

            reklama_svoya_amount = reklama_sales_responsibility_data.get('СВОЯ', 0)
            reklama_chuzhaya_amount = reklama_sales_responsibility_data.get('ЧУЖАЯ', 0)

            sheet_reklama['R70'] = float(reklama_svoya_amount) if reklama_svoya_amount else 0
            sheet_reklama['R71'] = float(reklama_chuzhaya_amount) if reklama_chuzhaya_amount else 0
        else:
            with client:
                ui.notify('Лист "Реклама" не найден в файле', type='warning')

        # 9. Обработка листа "Ангар"
        if progress_callback:
            progress_callback(90, "Обработка листа 'Ангар'...")
        if "Ангар" in wb.sheetnames:
            sheet_angar = wb["Ангар"]

            angar_monthly_cells = {
                'ПРОМЫШЛЕННОСТЬ': 'Q33', 'СЕЛЬСКОЕ ХОЗЯЙСТВО': 'Q34',
                'СПОРТ И КУЛЬТУРА': 'Q35', 'ПРОЧЕЕ': 'Q36'
            }

            for group_product, cell in angar_monthly_cells.items():
                amount = angar_monthly_group_data.get(group_product, 0)
                sheet_angar[cell] = float(amount) if amount else 0

            angar_sales_cells = {
                'ПРОМЫШЛЕННОСТЬ': 'Q6', 'СЕЛЬСКОЕ ХОЗЯЙСТВО': 'Q7',
                'СПОРТ И КУЛЬТУРА': 'Q8', 'ПРОЧЕЕ': 'Q9'
            }

            for group_product, cell in angar_sales_cells.items():
                amount = angar_sales_data.get(group_product, 0)
                sheet_angar[cell] = float(amount) if amount else 0
        else:
            with client:
                ui.notify('Лист "Ангар" не найден в файле', type='warning')

        # 10. Обработка листа "Ком"
        if progress_callback:
            progress_callback(95, "Обработка листа 'Ком'...")
        if "Ком" in wb.sheetnames:
            sheet_kn = wb["Ком"]

            kn_monthly_cells = {
                'ТКАНИ': 'R6', 'ФУРНИТУРА': 'R7', 'ПРОЧЕЕ': 'R8'
            }

            for group_product, cell in kn_monthly_cells.items():
                amount = kn_monthly_group_data.get(group_product, 0)
                sheet_kn[cell] = float(amount) if amount else 0

            kn_roznica_amount = kn_sales_data.get('РОЗНИЦА', 0)
            kn_potrebiteli_amount = kn_sales_data.get('ПОТРЕБИТЕЛИ', 0)

            sheet_kn['R29'] = float(kn_roznica_amount) if kn_roznica_amount else 0
            sheet_kn['R31'] = float(kn_potrebiteli_amount) if kn_potrebiteli_amount else 0

            kn_svoya_amount = kn_sales_responsibility_data.get('СВОЯ', 0)
            kn_chuzhaya_amount = kn_sales_responsibility_data.get('ЧУЖАЯ', 0)

            sheet_kn['R58'] = float(kn_svoya_amount) if kn_svoya_amount else 0
            sheet_kn['R59'] = float(kn_chuzhaya_amount) if kn_chuzhaya_amount else 0
        else:
            with client:
                ui.notify('Лист "Ком" не найден в файле', type='warning')

        # Сохраняем изменения
        if progress_callback:
            progress_callback(97, "Сохранение файла...")
        month_number = month_to_number[selected_month_name]
        output_filename = f"Самара {month_number} Отчет {selected_year} {selected_month_name}.xlsm"
        temp_dir = tempfile.gettempdir()
        output_path = os.path.join(temp_dir, output_filename)
        wb.save(output_path)
        wb.close()

        # Копируем значения из второго файла в первый
        if progress_callback:
            progress_callback(98, "Копирование данных из второго файла...")
        if file2_path and os.path.exists(file2_path):
            final_path = copy_cells_between_files(file2_path, output_path)
            if final_path:
                output_path = final_path

        if progress_callback:
            progress_callback(100, "Готово!")

        with client:
            ui.notify(f'Файл успешно создан: {os.path.basename(output_path)}', type='positive')
            ui.download(output_path)

    except Exception as e:
        with client:
            ui.notify(f'Ошибка при обработке файла: {e}', type='negative')
        print(traceback.format_exc())
    finally:
        # Удаляем временные файлы
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
            except:
                pass


async def handle_file1_upload(e):
    """Обработчик загрузки первого файла (шаблон)"""
    global file1_path

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsm') as tmp_file:
        file_content = await e.file.read()
        tmp_file.write(file_content)
        file1_path = tmp_file.name

    ui.notify(f'Файл шаблона загружен: {e.file.name}', type='positive')
    print(f"Файл шаблона сохранен: {file1_path}")


async def handle_file2_upload(e):
    """Обработчик загрузки второго файла (источник для копирования)"""
    global file2_path

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsm') as tmp_file:
        file_content = await e.file.read()
        tmp_file.write(file_content)
        file2_path = tmp_file.name

    ui.notify(f'Файл для копирования загружен: {e.file.name}', type='positive')
    print(f"Файл для копирования сохранен: {file2_path}")


def on_data_button_click():
    """Обработчик нажатия кнопки получения данных"""
    global cached_data, cached_sales_data, cached_monthly_group_data, cached_sales_responsibility_data
    global cached_tk_monthly_group_data, cached_tk_sales_responsibility_data, current_results_container
    global cached_reklama_monthly_group_data, cached_reklama_sales_responsibility_data, cached_reklama_total
    global cached_angar_monthly_group_data, cached_angar_sales_data
    global cached_kn_monthly_group_data, cached_kn_sales_data, cached_kn_sales_responsibility_data
    global cached_total_shipping_sum_angar, cached_total_shipping_sum_kn, cached_total_shipping_sum_reklama
    global cached_total_shipping_sum_tk

    selected_year = select_year.value
    selected_month_name = select_month.value
    selected_month_number = month_to_number[selected_month_name]

    ui.notify(f'Загружаю данные за {selected_month_name} {selected_year}...', type='info')

    data = fetch_data_by_direction(selected_year, selected_month_number)

    # Сохраняем данные для использования при загрузке файла
    directions_data = fetch_data_for_directions(selected_year, selected_month_number)
    sales_data = fetch_sales_data(selected_year, selected_month_number)
    monthly_group_data = fetch_monthly_group_products(selected_year, selected_month_number, 'ОАИ')
    sales_responsibility_data = fetch_sales_responsibility_data(selected_year, selected_month_number, 'ОАИ')

    tk_monthly_group_data = fetch_monthly_group_products(selected_year, selected_month_number, 'ТК')
    tk_sales_responsibility_data = fetch_sales_responsibility_data(selected_year, selected_month_number, 'ТК')

    reklama_monthly_group_data = fetch_monthly_group_products(selected_year, selected_month_number, 'РЕКЛАМА')
    reklama_sales_responsibility_data = fetch_sales_responsibility_data(selected_year, selected_month_number, 'РЕКЛАМА')
    reklama_total = fetch_reklama_total(selected_year, selected_month_number)

    angar_monthly_group_data = fetch_monthly_group_products(selected_year, selected_month_number, 'АНГАРЫ')
    angar_sales_data = fetch_angar_sales_data(selected_year, selected_month_number)

    kn_monthly_group_data = fetch_monthly_group_products(selected_year, selected_month_number, 'КН')
    kn_sales_data = fetch_kn_sales_data(selected_year, selected_month_number)
    kn_sales_responsibility_data = fetch_sales_responsibility_data(selected_year, selected_month_number, 'КН')

    total_shipping_angar = fetch_total_shipping_sum(selected_year, selected_month_number, 'АНГАРЫ')
    total_shipping_kn = fetch_total_shipping_sum(selected_year, selected_month_number, 'КН')
    total_shipping_reklama = fetch_total_shipping_sum(selected_year, selected_month_number, 'РЕКЛАМА')
    total_shipping_tk = fetch_total_shipping_sum(selected_year, selected_month_number, 'ТК')

    cached_data = directions_data
    cached_sales_data = sales_data
    cached_monthly_group_data = monthly_group_data
    cached_sales_responsibility_data = sales_responsibility_data
    cached_tk_monthly_group_data = tk_monthly_group_data
    cached_tk_sales_responsibility_data = tk_sales_responsibility_data
    cached_reklama_monthly_group_data = reklama_monthly_group_data
    cached_reklama_sales_responsibility_data = reklama_sales_responsibility_data
    cached_reklama_total = reklama_total
    cached_angar_monthly_group_data = angar_monthly_group_data
    cached_angar_sales_data = angar_sales_data
    cached_kn_monthly_group_data = kn_monthly_group_data
    cached_kn_sales_data = kn_sales_data
    cached_kn_sales_responsibility_data = kn_sales_responsibility_data
    cached_total_shipping_sum_angar = total_shipping_angar
    cached_total_shipping_sum_kn = total_shipping_kn
    cached_total_shipping_sum_reklama = total_shipping_reklama
    cached_total_shipping_sum_tk = total_shipping_tk

    # Очищаем и пересоздаем контейнер результатов
    if current_results_container:
        current_results_container.clear()

    with result_container:
        if data and len(data) > 0:
            ui.label(f'Результаты за {selected_month_name} {selected_year}:').classes('text-h6 mb-4')

            total_all = 0

            for direction, total_summ in data:
                if total_summ:
                    direction_total = float(total_summ)
                    total_all += direction_total
                    create_expandable_row(direction, direction_total, selected_year, selected_month_number)
                else:
                    create_expandable_row(direction, 0.0, selected_year, selected_month_number)

            ui.separator()
            with ui.row().classes('mt-4 justify-end w-full'):
                ui.label('Общая сумма по всем направлениям:').classes('text-h6 font-bold')
                ui.label(f'{total_all:,.2f} руб.').classes('text-h6 font-bold text-primary')
        else:
            ui.label(f'Нет данных за {selected_month_name} {selected_year}').classes('text-body1 text-grey-8')

    current_results_container = result_container


def on_process_button_click():
    """Обработчик нажатия кнопки обработки"""
    global file1_path, file2_path

    if not file1_path:
        ui.notify('Сначала загрузите файл шаблона', type='warning')
        return

    if not cached_data:
        ui.notify('Сначала получите данные из БД (кнопка "Получить данные")', type='warning')
        return

    selected_year = select_year.value
    selected_month_name = select_month.value

    # Получаем сумму реализации
    try:
        realization_amount_value = float(realization_input.value) if realization_input.value else 0
    except (ValueError, TypeError):
        realization_amount_value = 0

    # Сохраняем клиент из текущего контекста (основной поток)
    client = ui.context.client

    # Создаем диалог с прогресс-баром
    with ui.dialog() as dialog, ui.card():
        ui.label('Обработка файла...').classes('text-h6')
        progress_bar = ui.linear_progress(value=0, show_value=True).classes('w-full')
        progress_label = ui.label('Начало обработки...').classes('text-caption')

        def update_progress(value, message):
            progress_bar.value = value / 100
            progress_label.text = message

        dialog.open()

        # Запускаем обработку в отдельном потоке
        async def process():
            await asyncio.sleep(0.1)  # Даем время отобразиться диалогу

            process_excel_file(
                file1_path, file2_path, selected_month_name, selected_year,
                cached_data, cached_sales_data, cached_monthly_group_data,
                cached_sales_responsibility_data, cached_tk_monthly_group_data,
                cached_tk_sales_responsibility_data, cached_reklama_monthly_group_data,
                cached_reklama_sales_responsibility_data, cached_reklama_total,
                cached_angar_monthly_group_data, cached_angar_sales_data,
                cached_kn_monthly_group_data, cached_kn_sales_data,
                cached_kn_sales_responsibility_data,
                cached_total_shipping_sum_angar, cached_total_shipping_sum_kn,
                cached_total_shipping_sum_reklama, cached_total_shipping_sum_tk,
                realization_amount_value, client, update_progress
            )
            dialog.close()

        asyncio.create_task(process())


# Создаем интерфейс
with ui.card().classes('w-full p-4 mb-4'):
    ui.label('Настройки отчета').classes('text-h5 mb-2')

    with ui.row().classes('items-end gap-4'):
        select_year = ui.select(options=year, value=2026, label='Выберите год').classes('w-40')
        select_month = ui.select(options=month, value='Январь', label='Выберите месяц').classes('w-40')
        realization_input = ui.number(label='Сумма реализации (отгрузки)', value=0, step=1000, format='%.2f').classes(
            'w-48')
        ui.button('Получить данные', on_click=on_data_button_click, icon='search').classes('bg-primary text-white')

    ui.separator().classes('my-4')

    ui.label('Файлы для отчета').classes('text-h5 mb-2')

    with ui.row().classes('items-center gap-4'):
        ui.upload(label='Файл-шаблон (XLSM)', on_upload=handle_file1_upload, auto_upload=True).classes('w-auto').props(
            'accept=".xlsm"')
        ui.upload(label='Файл-источник (XLSM)', on_upload=handle_file2_upload, auto_upload=True).classes(
            'w-auto').props('accept=".xlsm"')
        ui.button('Создать отчет', on_click=on_process_button_click, icon='description', color='positive')

# Создаем пустой контейнер для результатов
result_container = ui.column().classes('w-full')

ui.run(title='Составление ежемесячного отчёта', reload=True, port=8001)